import modules.ir as ir
import modules.plot as plot
import modules.sinkhorn_algorithm as sink
import modules.colour_maps as cmap
import os
import numpy as np
import ot
import cv2
import scipy.spatial.distance as dist
from scipy.stats import chisquare, entropy
import openpyxl as xl
from openpyxl.styles import PatternFill, Font
from openpyxl.styles.borders import Border, Side
import inspect
import modules.comp_funcs as comp_funcs

		#####################
		#### SETUP START ####
		#####################


############################################ Comparison functions


funcs = [comp_funcs.wasserstein_distance, 
		 comp_funcs.wasserstein_distance_unbalanced, 
		 comp_funcs.l2,
		 comp_funcs.diagonality, 
		 comp_funcs.bhattacharyya, 
		 comp_funcs.correlation, 
		 comp_funcs.chi_square, 
		 comp_funcs.kl_divergence]

# funcs = [comp_funcs.l2, 
# 		 comp_funcs.diagonality, 
# 		 comp_funcs.bhattacharyya, 
# 		 comp_funcs.kl_divergence]

############################################ Colour map for data colouring


colour_map = cmap.WhiteGreen()


############################################ Directory with KF files


kf_dir = r"C:\Users\Yuman\Desktop\Programmeren\bachelorproject\scripting\RUNS\#KFFiles\second_set"

kf_dir = r"C:\Users\Yuman\Desktop\Programmeren\bachelorproject\scripting\RUNS\#KFFiles\random_set"


############################################ Categories to split the KF files into


# categories = ['Aminoindan_CONF_', 'ISO34_E', 'ISO34_P', 'SCONF_']
# categories = ['Aminoindan_CONF_']

categories = ['']


############################################ Path to save xl file to


save_name = kf_dir.split('\\')[-1] + '_comparisons2'
# save_name = kf_dir.split('\\')[-1] + '_comparisons'


#############################################



		#####################
		####  SETUP END  ####
		#####################




func_best = {comp_funcs.wasserstein_distance: 				min, 
			 comp_funcs.wasserstein_distance_unbalanced: 	min, 
			 comp_funcs.l2:									min, 
			 comp_funcs.diagonality: 						max, 
			 comp_funcs.bhattacharyya: 						min, 
			 comp_funcs.correlation:						min, 
			 comp_funcs.chi_square: 						min, 
			 comp_funcs.kl_divergence: 						min}


#get kffiles
kff = [f for f in os.listdir(kf_dir)]

#separate into categories
kff_by_cat = {}
for cat in categories:
	kff_by_cat[cat] = [kf_dir + '\\' + f for f in kff if f.startswith(cat)]

names_by_cat = {}
for cat in categories:
	names_by_cat[cat] = []
	for f in kff:
		name = f
		if name.startswith(cat) and name.endswith('.t21'):
			name = name[len(cat):][:-8]
			names_by_cat[cat].append(name)





#start writing
wb = xl.Workbook()
ws0 = wb.active

bottom_line = Border(bottom=Side(style='thin'))
right_line = Border(right=Side(style='thin'))
both_line = Border(right=Side(style='thin'), bottom=Side(style='thin'))
all_line = Border(top=Side(style='thick'), left=Side(style='thick'), right=Side(style='thick'), bottom=Side(style='thick'))


bold_huge = Font(bold=False, size=30)
bold_large = Font(bold=False, size=16)
bold_small = Font(bold=False, size=11)



for cat in categories:

	#get the spectra
	ir_dft = [ir.get_spectrum_from_kf(f,width=50) for f in kff_by_cat[cat] if f.endswith('.t21')]
	ir_dftb = [ir.get_spectrum_from_kf(f,width=50) for f in kff_by_cat[cat] if f.endswith('.rkf')]


	# print(np.sum(ir_dft), np.sum(ir_dftb))


	#setup sheet
	ws = wb.create_sheet(cat)
	wb.active = ws
	ws['A1'] = cat
	ws['A1'].font = bold_huge

	data_col_offset = 5

	n = len(ir_dft)
	for i, func in enumerate(funcs):

		#setup new matching table
		fn = str(func).split()[1] #func name
		print(cat + '_' + fn)

		ws.cell(row=i*(n+3)+3, column=1).value = fn.upper()
		ws.cell(row=i*(n+3)+3, column=1).font = bold_large

		ws.cell(row=i*(n+3)+4, column=1).value = 'Best'
		ws.cell(row=i*(n+3)+4, column=1).border = bottom_line
		ws.cell(row=i*(n+3)+4, column=1).font = bold_small

		ws.cell(row=i*(n+3)+4, column=2).value = 'Error'
		ws.cell(row=i*(n+3)+4, column=2).border = bottom_line
		ws.cell(row=i*(n+3)+4, column=2).font = bold_small

		ws.cell(row=i*(n+3)+4, column=3).value = 'Error rel'
		ws.cell(row=i*(n+3)+4, column=3).border = bottom_line
		ws.cell(row=i*(n+3)+4, column=3).font = bold_small

		ws.cell(row=i*(n+3)+4, column=data_col_offset).value = r'DFT\DFTB'
		ws.cell(row=i*(n+3)+4, column=data_col_offset).border = both_line
		ws.cell(row=i*(n+3)+4, column=data_col_offset).font = bold_small

		# print(type(ir_dft), type(ir_dftb))
		d = func(ir_dft, ir_dftb)
		print(d)
		print(d.max(), d.min())
		
		if func_best[func] is min:
			dnorm = (d-d.min())/(d.max()-d.min())
		if func_best[func] is max:
			dnorm = 1-(d-d.min())/(d.max()-d.min())

		c = colour_map.get_hex_colours(dnorm)

		for j, row in enumerate(d):
			try:
				ws.cell(row=i*(n+3)+5+j, column=data_col_offset).value = names_by_cat[cat][j]
				ws.cell(row=i*(n+3)+5+j, column=data_col_offset).border = right_line

				ws.cell(row=i*(n+3)+5+j, column=1).value = func_best[func](row)
				ws.cell(row=i*(n+3)+5+j, column=2).value = row[j] - func_best[func](row)
				ws.cell(row=i*(n+3)+5+j, column=3).value = f'{(row[j] - func_best[func](row))/row[j] * 100:.2f}%'


				for k, el in enumerate(row):
					ws.cell(row=i*(n+3)+4, column=data_col_offset+1+k).value = names_by_cat[cat][k]
					ws.cell(row=i*(n+3)+4, column=data_col_offset+1+k).border = bottom_line

					ws.cell(row=i*(n+3)+5+j, column=data_col_offset+1+k).value = el
					ws.cell(row=i*(n+3)+5+j, column=data_col_offset+1+k).fill = PatternFill(start_color=c[j,k].decode("utf-8"), end_color=c[j,k].decode("utf-8"), fill_type = "solid")

					if el == func_best[func](row):
						ws.cell(row=i*(n+3)+5+j, column=data_col_offset+1+k).border = all_line
			except:
				pass

		wb.save(f"{save_name}.xlsx")


del wb['Sheet']


wb.save(f"{save_name}.xlsx")