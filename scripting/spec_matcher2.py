import modules.ir as ir
import modules.plot as plot
import modules.sinkhorn_algorithm as sink
import modules.colour_maps as cmap
import os
import numpy as np
import ot
import cv2
import scipy.spatial.distance as dist
from scipy.stats import chisquare, entropy
import openpyxl as xl
from openpyxl.styles import PatternFill, Font
from openpyxl.styles.borders import Border, Side
import inspect
import modules.comp_funcs as comp_funcs



def main(functionals, bins=600):

			#####################
			#### SETUP START ####
			#####################


	############################################ Comparison functions

	funcs = [comp_funcs.wasserstein_distance, 
			 comp_funcs.wasserstein_distance_unbalanced, 
			 comp_funcs.l2,
			 comp_funcs.diagonality, 
			 comp_funcs.bhattacharyya, 
			 comp_funcs.correlation, 
			 comp_funcs.chi_square, 
			 comp_funcs.kl_divergence]

	# funcs = [comp_funcs.wasserstein_distance_unbalanced]

	############################################ Colour map for data colouring


	colour_map = cmap.WhiteGreen()


	############################################ Directory with KF files


	kf_dir = r"D:\Users\Yuman\Desktop\Programmeren\bachelorproject\scripting\RUNS\#KFFiles\functionals"

	# kf_dir = r"C:\Users\Yuman\Desktop\Programmeren\bachelorproject\scripting\RUNS\#KFFiles\random_set"


	############################################ Categories to split the KF files into


	# categories = ['Aminoindan_CONF_', 'ISO34_E', 'ISO34_P', 'SCONF']
	# categories = ['Aminoindan_CONF_', 'ISO34_E', 'SCONF']
	categories = ['Aminoindan_CONF_']


	############################################ Sheet names


	sheet_names = categories


	############################################ Functionals to split the KF files into


	# functionals = ['OLYP_DFT', 'LDA_DFT', 'DFTB3_DFTB', 'DFTB3_freq_DFTB']
	# functionals = ['DFTB3_freq_DFTB', 'LDA_DFT']


	############################################ Path to save xl file to


	save_name = kf_dir.split('\\')[-1] + f'_comparisons_{functionals[0]}_{functionals[1]}'
	# save_name = kf_dir.split('\\')[-1] + '_comparisons'


	#############################################





			#####################
			####  SETUP END  ####
			#####################


	assert len(categories) == len(sheet_names)



	func_best = {comp_funcs.wasserstein_distance: 				min, 
				 comp_funcs.wasserstein_distance_unbalanced: 	min, 
				 comp_funcs.l2:									min, 
				 comp_funcs.diagonality: 						max, 
				 comp_funcs.bhattacharyya: 						min, 
				 comp_funcs.correlation:						min, 
				 comp_funcs.chi_square: 						min, 
				 comp_funcs.kl_divergence: 						min}


	#get kffiles
	kff = [f for f in os.listdir(kf_dir)]
	kff = list(filter(lambda x: functionals[0] in x or functionals[1] in x, kff))

	#separate into categories
	kff_by_cat = {}
	for cat in categories:
		kff_by_cat[cat] = [kf_dir + '\\' + f for f in kff if f.startswith(cat)]

	names_by_cat = {}
	for cat in categories:
		names_by_cat[cat] = {functionals[0]:[], functionals[1]:[]}
		for f in kff:
			name = f
			if name.startswith(cat):
				name = name[len(cat):][:-4]
				if functionals[0] in name:
					names_by_cat[cat][functionals[0]].append(name)
				elif functionals[1] in name:
					names_by_cat[cat][functionals[1]].append(name)




	#start writing
	out_file = f"{save_name}_{bins}.xlsx"
	wb = xl.Workbook()
	ws0 = wb.active

	bottom_line = Border(bottom=Side(style='thin'))
	right_line = Border(right=Side(style='thin'))
	both_line = Border(right=Side(style='thin'), bottom=Side(style='thin'))
	all_line = Border(top=Side(style='thick'), left=Side(style='thick'), right=Side(style='thick'), bottom=Side(style='thick'))


	bold_huge = Font(bold=False, size=30)
	bold_large = Font(bold=False, size=16)
	bold_small = Font(bold=False, size=11)



	for i_cat, s_name, cat in zip(range(len(categories)), sheet_names, categories):

		#get the spectra
		ir1 = [ir.get_spectrum_from_kf(f,width=50,n=bins) for f in kff_by_cat[cat] if functionals[0] in f]
		ir2 = [ir.get_spectrum_from_kf(f,width=50,n=bins) for f in kff_by_cat[cat] if functionals[1] in f]


		#setup sheet
		ws = wb.create_sheet(s_name)
		wb.active = ws
		ws['A1'] = s_name
		ws['A1'].font = bold_huge

		data_col_offset = 5

		n = len(ir1)
		for i, func in enumerate(funcs):

			#setup new matching table
			fn = str(func).split()[1] #func name
			print(cat + '_' + fn)

			ws.cell(row=i*(n+3)+3, column=1).value = fn.upper()
			ws.cell(row=i*(n+3)+3, column=1).font = bold_large

			ws.cell(row=i*(n+3)+4, column=1).value = 'Best'
			ws.cell(row=i*(n+3)+4, column=1).border = bottom_line
			ws.cell(row=i*(n+3)+4, column=1).font = bold_small

			ws.cell(row=i*(n+3)+4, column=2).value = 'Error'
			ws.cell(row=i*(n+3)+4, column=2).border = bottom_line
			ws.cell(row=i*(n+3)+4, column=2).font = bold_small

			ws.cell(row=i*(n+3)+4, column=3).value = 'Error rel'
			ws.cell(row=i*(n+3)+4, column=3).border = bottom_line
			ws.cell(row=i*(n+3)+4, column=3).font = bold_small

			ws.cell(row=i*(n+3)+4, column=data_col_offset).value = rf'{functionals[0]}\{functionals[1]}'
			ws.cell(row=i*(n+3)+4, column=data_col_offset).border = both_line
			ws.cell(row=i*(n+3)+4, column=data_col_offset).font = bold_small

			# kwargs = {'reg_m': np.linspace(-1,2,7)[i_cat]}

			d = func(ir1, ir2, reg_m=10**2)
			

			if func_best[func] is min:
				dnorm = (d-d.min())/(d.max()-d.min())
			if func_best[func] is max:
				dnorm = 1-(d-d.min())/(d.max()-d.min())

			c = colour_map.get_hex_colours(dnorm)

			for j, row in enumerate(d):
				try:
					ws.cell(row=i*(n+3)+5+j, column=data_col_offset).value = names_by_cat[cat][functionals[0]][j]
					ws.cell(row=i*(n+3)+5+j, column=data_col_offset).border = right_line

					ws.cell(row=i*(n+3)+5+j, column=1).value = func_best[func](row)
					ws.cell(row=i*(n+3)+5+j, column=2).value = row[j] - func_best[func](row)
					ws.cell(row=i*(n+3)+5+j, column=3).value = f'{(row[j] - func_best[func](row))/row[j] * 100:.2f}%'


					for k, el in enumerate(row):
						ws.cell(row=i*(n+3)+4, column=data_col_offset+1+k).value = names_by_cat[cat][functionals[1]][k]
						ws.cell(row=i*(n+3)+4, column=data_col_offset+1+k).border = bottom_line

						ws.cell(row=i*(n+3)+5+j, column=data_col_offset+1+k).value = el
						ws.cell(row=i*(n+3)+5+j, column=data_col_offset+1+k).fill = PatternFill(start_color=c[j,k].decode("utf-8"), end_color=c[j,k].decode("utf-8"), fill_type = "solid")

						if el == func_best[func](row):
							ws.cell(row=i*(n+3)+5+j, column=data_col_offset+1+k).border = all_line
				except:
					raise

			wb.save(out_file)


	del wb['Sheet']
	wb.save(out_file)



main(['LDA_DFT','DFTB3_DFTB'],bins=600)
# main(['LDA_DFT','DFTB3_DFTB'],bins=2400)
# main(['OLYP_DFT','DFTB3_DFTB'])
# main(['OLYP_DFT','LDA_DFT'])